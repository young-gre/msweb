import os
import re
import io
import sqlite3
import secrets
import threading
import time
import pandas as pd
from datetime import datetime
from flask import (Flask, request, jsonify, render_template,
                   send_file, session, redirect, url_for)
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'msweb-secret-key-chungbuk-2026'

BASE_DIR        = os.path.dirname(__file__)
DB_PATH         = os.path.join(BASE_DIR, 'ms_data.db')
UPLOAD_PASSWORD = '1234'

CHEONGJU_DISTS = {'청주시 상당구','청주시 서원구','청주시 청원구','청주시 흥덕구'}

CAR_CLASS_GROUP = {
    '경승용'   : ('승용', '경승용'),
    '소형'     : ('승용', '소형'),
    '준중형'   : ('승용', '준중형'),
    '중형'     : ('승용', '중형'),
    '대형'     : ('승용', '대형'),
    '스포츠'   : ('승용', '스포츠'),
    '경SUV'    : ('RV',   '경SUV'),
    '소형SUV'  : ('RV',   '소형SUV'),
    '준중형SUV': ('RV',   '준중형SUV'),
    '중형SUV'  : ('RV',   '중형SUV'),
    '대형SUV'  : ('RV',   '대형SUV'),
    '소형MPV'  : ('RV',   'MPV'),
    '중대형MPV': ('RV',   'MPV'),
    '소형버스' : ('상용', '소형버스'),
    '소형트럭' : ('상용', '소형트럭'),
    '경상용'   : ('상용', '소형트럭'),
}

MAKER_GROUP = {
    '현대'    : '현대',
    '기아'    : '기아',
    'TESLA'   : '테슬라',
    'BYD'     : 'BYD',
    '르노삼성': '르노',
}

SECTION_ORDER = ['승용', 'RV', '상용']
CLASS_ORDER = {
    '승용': ['경승용', '소형', '준중형', '중형', '대형', '스포츠'],
    'RV'  : ['경SUV', '소형SUV', '준중형SUV', '중형SUV', '대형SUV', 'MPV'],
    '상용': ['소형버스', '소형트럭'],
}

MAKER_NAMES   = ['현대', '기아', '테슬라', 'BYD', '르노']
EV_ONLY_FUELS = {'전기'}

# ═══════════════════════════════════════════════════════
# DB 커넥션 풀 (스레드 로컬)
# ═══════════════════════════════════════════════════════
_local = threading.local()

def get_con():
    if not hasattr(_local, 'con') or _local.con is None:
        con = sqlite3.connect(DB_PATH, check_same_thread=False)
        con.row_factory = sqlite3.Row
        con.execute('PRAGMA journal_mode=WAL')
        con.execute('PRAGMA cache_size=20000')
        con.execute('PRAGMA synchronous=NORMAL')
        con.execute('PRAGMA temp_store=MEMORY')
        con.execute('PRAGMA mmap_size=268435456')
        _local.con = con
    return _local.con

def query_db(sql, params=()):
    con  = get_con()
    rows = con.execute(sql, params).fetchall()
    return [dict(r) for r in rows]

# ═══════════════════════════════════════════════════════
# 전국 M/S 캐시
# ═══════════════════════════════════════════════════════
_nat_cache      = {}
_nat_cache_time = {}
NAT_CACHE_TTL   = 300

# ═══════════════════════════════════════════════════════
# 집계 준비 완료 플래그
# ═══════════════════════════════════════════════════════
_agg_ready = False
_agg_lock  = threading.Lock()

# ═══════════════════════════════════════════════════════
# 헬퍼 함수
# ═══════════════════════════════════════════════════════
def classify_maker(maker):
    return MAKER_GROUP.get(maker, '외산차')

def get_maker_col(maker, import_country):
    if maker == '현대':     return '현대'
    if maker == '기아':     return '기아'
    if maker == 'TESLA':    return '테슬라'
    if maker == 'BYD':      return 'BYD'
    if maker == '르노삼성': return '르노'
    if import_country != '대한민국': return '외산차'
    return None

def get_brand_label(maker, import_country):
    if maker == '현대':     return '현대'
    if maker == '기아':     return '기아'
    if maker == 'TESLA':    return 'TESLA'
    if maker == 'BYD':      return 'BYD'
    if maker == '르노삼성': return '르노삼성'
    if import_country == '대한민국': return '국산기타'
    return maker

def get_table(mode):
    if mode == 'base':     return 'car_reg_base'
    if mode == 'national': return 'national_reg'
    return 'car_reg'

def get_agg_table(mode, accum=False):
    if accum:
        if mode == 'base':     return 'agg_base_accum'
        if mode == 'national': return 'agg_national_accum'
        return 'agg_sigungu_accum'
    else:
        if mode == 'base':     return 'agg_base'
        if mode == 'national': return 'agg_national'
        return 'agg_sigungu'

def get_tbl_for_mode(mode, args):
    accum = args.get('accum', '') == '1'
    if mode == 'base':
        base_sg   = args.get('base_sigungu', '').strip()
        base_dong = args.get('base_dong',    '').strip()
        if (base_sg and base_sg != '전체') or (base_dong and base_dong != '전체'):
            return 'car_reg'
    return get_agg_table(mode, accum)

def fuel_group_clause(fuel_group):
    if fuel_group == 'ice':
        return ["fuel NOT LIKE '%전기%'","fuel NOT LIKE '%수소%'","fuel NOT LIKE '%하이브리드%'"]
    elif fuel_group == 'hev':
        return ["fuel LIKE '%하이브리드%'"]
    elif fuel_group == 'ev':
        return ["fuel LIKE '%전기%'","fuel NOT LIKE '%수소%'","fuel NOT LIKE '%하이브리드%'"]
    elif fuel_group == 'fcev':
        return ["fuel LIKE '%수소%'"]
    return []

def get_orig_classes(sec=None, cls=None):
    return [cc for cc, (s, c) in CAR_CLASS_GROUP.items()
            if (not sec or s == sec) and (not cls or c == cls)]

def add_loc_clause(clauses, params, loc_col, loc_val):
    if loc_val == '청주시':
        ph = ','.join(['?' for _ in CHEONGJU_DISTS])
        clauses.append(f'{loc_col} IN ({ph})')
        params.extend(sorted(CHEONGJU_DISTS))
    else:
        clauses.append(f'{loc_col} = ?')
        params.append(loc_val)

# ═══════════════════════════════════════════════════════
# DB 초기화
# ═══════════════════════════════════════════════════════
def init_db():
    with sqlite3.connect(DB_PATH) as con:
        con.execute('PRAGMA journal_mode=WAL')
        con.execute('PRAGMA cache_size=20000')
        con.execute('PRAGMA synchronous=NORMAL')
        con.execute('PRAGMA temp_store=MEMORY')

        con.execute('''CREATE TABLE IF NOT EXISTS car_reg (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            headquarters TEXT, region TEXT, sigungu TEXT, dong TEXT,
            car_class TEXT, import_country TEXT, maker TEXT, fuel TEXT, model TEXT,
            base TEXT, year INTEGER, month INTEGER, reg_count INTEGER)''')

        existing = [r[1] for r in con.execute('PRAGMA table_info(car_reg)').fetchall()]
        for col in ['dong','base','headquarters']:
            if col not in existing:
                con.execute(f'ALTER TABLE car_reg ADD COLUMN {col} TEXT DEFAULT ""')

        con.execute('''CREATE TABLE IF NOT EXISTS car_reg_base (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            headquarters TEXT, base TEXT, car_class TEXT, import_country TEXT,
            maker TEXT, fuel TEXT, model TEXT, year INTEGER, month INTEGER, reg_count INTEGER)''')

        existing_base = [r[1] for r in con.execute('PRAGMA table_info(car_reg_base)').fetchall()]
        if 'headquarters' not in existing_base:
            con.execute('ALTER TABLE car_reg_base ADD COLUMN headquarters TEXT DEFAULT ""')

        con.execute('''CREATE TABLE IF NOT EXISTS national_reg (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            car_class TEXT, import_country TEXT, maker TEXT, fuel TEXT, model TEXT,
            year INTEGER, month INTEGER, reg_count INTEGER)''')

        con.execute('''CREATE TABLE IF NOT EXISTS upload_meta (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mode TEXT, year INTEGER, month INTEGER, uploaded_at TEXT)''')

        con.execute('''CREATE TABLE IF NOT EXISTS data_memo (
            id INTEGER PRIMARY KEY, mode TEXT UNIQUE, memo TEXT, saved_at TEXT)''')

        idxs = [
            'CREATE INDEX IF NOT EXISTS idx_cr_ym      ON car_reg(year,month)',
            'CREATE INDEX IF NOT EXISTS idx_cr_ym_sg   ON car_reg(year,month,sigungu)',
            'CREATE INDEX IF NOT EXISTS idx_cr_ym_hq   ON car_reg(year,month,headquarters)',
            'CREATE INDEX IF NOT EXISTS idx_cr_ym_mk   ON car_reg(year,month,maker)',
            'CREATE INDEX IF NOT EXISTS idx_cr_base    ON car_reg(base)',
            'CREATE INDEX IF NOT EXISTS idx_crb_ym     ON car_reg_base(year,month)',
            'CREATE INDEX IF NOT EXISTS idx_crb_ym_base ON car_reg_base(year,month,base)',
            'CREATE INDEX IF NOT EXISTS idx_crb_ym_hq  ON car_reg_base(year,month,headquarters)',
            'CREATE INDEX IF NOT EXISTS idx_nat_ym     ON national_reg(year,month)',
            'CREATE INDEX IF NOT EXISTS idx_nat_ym_mk  ON national_reg(year,month,maker)',
        ]
        for idx in idxs:
            con.execute(idx)

        con.execute('''CREATE TABLE IF NOT EXISTS agg_sigungu (
            year INTEGER, month INTEGER,
            headquarters TEXT, sigungu TEXT,
            car_class TEXT, maker TEXT, import_country TEXT, fuel TEXT, model TEXT,
            reg_count INTEGER)''')
        con.execute('''CREATE TABLE IF NOT EXISTS agg_base (
            year INTEGER, month INTEGER,
            headquarters TEXT, base TEXT,
            car_class TEXT, maker TEXT, import_country TEXT, fuel TEXT, model TEXT,
            reg_count INTEGER)''')
        con.execute('''CREATE TABLE IF NOT EXISTS agg_national (
            year INTEGER, month INTEGER,
            car_class TEXT, maker TEXT, fuel TEXT, model TEXT,
            reg_count INTEGER)''')
        con.execute('''CREATE TABLE IF NOT EXISTS agg_sigungu_accum (
            year INTEGER, month INTEGER,
            headquarters TEXT, sigungu TEXT,
            car_class TEXT, maker TEXT, import_country TEXT, fuel TEXT, model TEXT,
            reg_count INTEGER)''')
        con.execute('''CREATE TABLE IF NOT EXISTS agg_base_accum (
            year INTEGER, month INTEGER,
            headquarters TEXT, base TEXT,
            car_class TEXT, maker TEXT, import_country TEXT, fuel TEXT, model TEXT,
            reg_count INTEGER)''')
        con.execute('''CREATE TABLE IF NOT EXISTS agg_national_accum (
            year INTEGER, month INTEGER,
            car_class TEXT, maker TEXT, fuel TEXT, model TEXT,
            reg_count INTEGER)''')

        agg_idxs = [
            'CREATE INDEX IF NOT EXISTS idx_agg_sg_ym     ON agg_sigungu(year,month)',
            'CREATE INDEX IF NOT EXISTS idx_agg_sg_ym_hq  ON agg_sigungu(year,month,headquarters)',
            'CREATE INDEX IF NOT EXISTS idx_agg_sg_ym_loc ON agg_sigungu(year,month,sigungu)',
            'CREATE INDEX IF NOT EXISTS idx_agg_sg_mk     ON agg_sigungu(year,month,maker)',
            'CREATE INDEX IF NOT EXISTS idx_agg_ba_ym     ON agg_base(year,month)',
            'CREATE INDEX IF NOT EXISTS idx_agg_ba_ym_hq  ON agg_base(year,month,headquarters)',
            'CREATE INDEX IF NOT EXISTS idx_agg_ba_ym_loc ON agg_base(year,month,base)',
            'CREATE INDEX IF NOT EXISTS idx_agg_ba_mk     ON agg_base(year,month,maker)',
            'CREATE INDEX IF NOT EXISTS idx_agg_nat_ym    ON agg_national(year,month)',
            'CREATE INDEX IF NOT EXISTS idx_agg_nat_mk    ON agg_national(year,month,maker)',
            'CREATE INDEX IF NOT EXISTS idx_agg_sgac_ym   ON agg_sigungu_accum(year,month)',
            'CREATE INDEX IF NOT EXISTS idx_agg_sgac_hq   ON agg_sigungu_accum(year,month,headquarters)',
            'CREATE INDEX IF NOT EXISTS idx_agg_sgac_loc  ON agg_sigungu_accum(year,month,sigungu)',
            'CREATE INDEX IF NOT EXISTS idx_agg_baac_ym   ON agg_base_accum(year,month)',
            'CREATE INDEX IF NOT EXISTS idx_agg_baac_hq   ON agg_base_accum(year,month,headquarters)',
            'CREATE INDEX IF NOT EXISTS idx_agg_baac_loc  ON agg_base_accum(year,month,base)',
            'CREATE INDEX IF NOT EXISTS idx_agg_natac_ym  ON agg_national_accum(year,month)',
        ]
        for idx in agg_idxs:
            con.execute(idx)
        con.commit()

# ═══════════════════════════════════════════════════════
# 사전 집계 빌드
# ═══════════════════════════════════════════════════════
def _build_monthly_agg(con, years_months=None):
    def ym_cond():
        if years_months:
            ph = ','.join(['(?,?)'] * len(years_months))
            flat = [v for ym in years_months for v in ym]
            return f'(year,month) IN ({ph})', flat
        return '1=1', []

    cond, flat = ym_cond()

    if years_months:
        for y, m in years_months:
            con.execute('DELETE FROM agg_sigungu  WHERE year=? AND month=?', (y,m))
            con.execute('DELETE FROM agg_base     WHERE year=? AND month=?', (y,m))
            con.execute('DELETE FROM agg_national WHERE year=? AND month=?', (y,m))
    else:
        con.execute('DELETE FROM agg_sigungu')
        con.execute('DELETE FROM agg_base')
        con.execute('DELETE FROM agg_national')

    con.execute(f'''
        INSERT INTO agg_sigungu
            (year,month,headquarters,sigungu,car_class,maker,import_country,fuel,model,reg_count)
        SELECT year,month,
            COALESCE(headquarters,''), COALESCE(sigungu,''),
            COALESCE(car_class,''), COALESCE(maker,''),
            COALESCE(import_country,''), COALESCE(fuel,''),
            COALESCE(model,''), SUM(reg_count)
        FROM car_reg
        WHERE sigungu!='' AND {cond}
        GROUP BY year,month,headquarters,sigungu,car_class,maker,import_country,fuel,model
    ''', flat)

    con.execute(f'''
        INSERT INTO agg_base
            (year,month,headquarters,base,car_class,maker,import_country,fuel,model,reg_count)
        SELECT year,month,
            COALESCE(headquarters,''), COALESCE(base,''),
            COALESCE(car_class,''), COALESCE(maker,''),
            COALESCE(import_country,''), COALESCE(fuel,''),
            COALESCE(model,''), SUM(reg_count)
        FROM car_reg_base
        WHERE base!='' AND {cond}
        GROUP BY year,month,headquarters,base,car_class,maker,import_country,fuel,model
    ''', flat)

    con.execute(f'''
        INSERT INTO agg_national
            (year,month,car_class,maker,fuel,model,reg_count)
        SELECT year,month,
            COALESCE(car_class,''), COALESCE(maker,''),
            COALESCE(fuel,''), COALESCE(model,''), SUM(reg_count)
        FROM national_reg
        WHERE {cond}
        GROUP BY year,month,car_class,maker,fuel,model
    ''', flat)


def _build_accum_agg(con):
    con.execute('DELETE FROM agg_sigungu_accum')
    con.execute('DELETE FROM agg_base_accum')
    con.execute('DELETE FROM agg_national_accum')

    ym_list = con.execute(
        'SELECT DISTINCT year, month FROM agg_sigungu ORDER BY year, month'
    ).fetchall()
    for year, month in ym_list:
        con.execute('''
            INSERT INTO agg_sigungu_accum
                (year,month,headquarters,sigungu,car_class,maker,import_country,fuel,model,reg_count)
            SELECT ?,?,
                COALESCE(headquarters,''), COALESCE(sigungu,''),
                COALESCE(car_class,''), COALESCE(maker,''),
                COALESCE(import_country,''), COALESCE(fuel,''),
                COALESCE(model,''), SUM(reg_count)
            FROM car_reg
            WHERE year=? AND month<=? AND sigungu!=''
            GROUP BY headquarters,sigungu,car_class,maker,import_country,fuel,model
        ''', (year, month, year, month))

    con.commit()

    ym_list_base = con.execute(
        'SELECT DISTINCT year, month FROM agg_base ORDER BY year, month'
    ).fetchall()
    for year, month in ym_list_base:
        con.execute('''
            INSERT INTO agg_base_accum
                (year,month,headquarters,base,car_class,maker,import_country,fuel,model,reg_count)
            SELECT ?,?,
                COALESCE(headquarters,''), COALESCE(base,''),
                COALESCE(car_class,''), COALESCE(maker,''),
                COALESCE(import_country,''), COALESCE(fuel,''),
                COALESCE(model,''), SUM(reg_count)
            FROM car_reg_base
            WHERE year=? AND month<=? AND base!=''
            GROUP BY headquarters,base,car_class,maker,import_country,fuel,model
        ''', (year, month, year, month))

    con.commit()

    ym_list_nat = con.execute(
        'SELECT DISTINCT year, month FROM agg_national ORDER BY year, month'
    ).fetchall()
    for year, month in ym_list_nat:
        con.execute('''
            INSERT INTO agg_national_accum
                (year,month,car_class,maker,fuel,model,reg_count)
            SELECT ?,?,
                COALESCE(car_class,''), COALESCE(maker,''),
                COALESCE(fuel,''), COALESCE(model,''), SUM(reg_count)
            FROM national_reg
            WHERE year=? AND month<=?
            GROUP BY car_class,maker,fuel,model
        ''', (year, month, year, month))

    con.commit()


def build_all_agg(years_months=None):
    global _agg_ready
    with _agg_lock:
        t0  = time.time()
        con = sqlite3.connect(DB_PATH)
        con.execute('PRAGMA journal_mode=WAL')
        con.execute('PRAGMA synchronous=NORMAL')
        con.execute('PRAGMA cache_size=20000')
        con.execute('PRAGMA temp_store=MEMORY')
        try:
            _build_monthly_agg(con, years_months)
            _build_accum_agg(con)
            con.commit()
            _nat_cache.clear()
            _nat_cache_time.clear()
            _agg_ready = True
            print(f'[AGG] 집계 완료 ({time.time()-t0:.2f}s)')
        except Exception as e:
            con.rollback()
            print(f'[AGG] 오류: {e}')
            raise
        finally:
            con.close()

# ═══════════════════════════════════════════════════════
# 전국 M/S 캐시
# ═══════════════════════════════════════════════════════
def get_national_ms(year, month, accum=False):
    if not year or not month: return {}, 0
    cache_key = (year, month, accum)
    now       = time.time()
    if cache_key in _nat_cache and now - _nat_cache_time.get(cache_key,0) < NAT_CACHE_TTL:
        return _nat_cache[cache_key]

    tbl = 'agg_national_accum' if accum else 'agg_national'
    nat_grand = query_db(
        f'SELECT COALESCE(SUM(reg_count),0) AS v FROM {tbl} WHERE year=? AND month=?',
        [year, month])[0]['v']

    if not nat_grand:
        result = {}, 0
    else:
        rows = query_db(
            f'SELECT maker, SUM(reg_count) AS total FROM {tbl} '
            f'WHERE year=? AND month=? GROUP BY maker', [year, month])
        result = ({r['maker']: round(r['total']/nat_grand*100,1) for r in rows}, nat_grand)

    _nat_cache[cache_key]      = result
    _nat_cache_time[cache_key] = now
    return result

# ═══════════════════════════════════════════════════════
# 업로드 / 파싱
# ═══════════════════════════════════════════════════════
def record_upload_meta(mode, years, months):
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    con = get_con()
    for y in years:
        for m in months:
            con.execute('DELETE FROM upload_meta WHERE mode=? AND year=? AND month=?', (mode,y,m))
            con.execute('INSERT INTO upload_meta(mode,year,month,uploaded_at) VALUES(?,?,?,?)',
                        (mode,y,m,now))
    con.commit()


def parse_and_insert_unified(filepath, filter_year=None, filter_month=None):
    df = pd.read_excel(filepath, header=0)
    df.columns = df.columns.astype(str).str.strip()

    date_cols = [c for c in df.columns if re.match(r'^\d{4}\.\d{2}$', str(c).strip())]
    id_cols   = [c for c in df.columns if c not in date_cols]
    if not date_cols:
        raise ValueError("날짜 형식 컬럼(YYYY.MM)을 찾을 수 없습니다.")

    if filter_year and filter_month:
        target = f"{filter_year}.{int(filter_month):02d}"
        if target not in date_cols:
            raise ValueError(f"{target} 컬럼이 없습니다.")
        date_cols = [target]

    long_df = df.melt(id_vars=id_cols, value_vars=date_cols,
                      var_name='yearmonth', value_name='reg_count')
    long_df = long_df[long_df['reg_count'].notna() & (long_df['reg_count'] > 0)].copy()
    long_df['reg_count'] = long_df['reg_count'].astype(int)
    long_df['year']  = long_df['yearmonth'].str[:4].astype(int)
    long_df['month'] = long_df['yearmonth'].str[5:7].astype(int)
    long_df.drop(columns=['yearmonth'], inplace=True)

    COL_MAP = {
        '지역본부':'headquarters','거점':'base','지역':'region',
        '시군구':'sigungu','법정동':'dong','등록차급':'car_class',
        '수입국가':'import_country','제조사':'maker','등록연료':'fuel','대표차명':'model',
    }
    long_df.rename(columns=COL_MAP, inplace=True)

    for col in ['headquarters','base','region','sigungu','dong',
                'car_class','import_country','maker','fuel','model']:
        long_df[col] = long_df.get(col, pd.Series([''] * len(long_df))) \
                               .fillna('').astype(str).str.strip()

    long_df = long_df[long_df['maker'] != '']
    years   = long_df['year'].unique().tolist()
    months  = long_df['month'].unique().tolist()
    ym_list = [(int(y), int(m)) for y in years for m in months]

    sg_df   = long_df[long_df['sigungu'] != ''].copy()
    base_df = long_df[long_df['base']    != ''].copy()

    con = sqlite3.connect(DB_PATH)
    try:
        for y, m in ym_list:
            con.execute('DELETE FROM car_reg      WHERE year=? AND month=?', (y,m))
            con.execute('DELETE FROM car_reg_base WHERE year=? AND month=?', (y,m))
        con.commit()
        sg_df.to_sql('car_reg',      con, if_exists='append', index=False)
        base_df.to_sql('car_reg_base', con, if_exists='append', index=False)
        con.commit()
    finally:
        con.close()

    record_upload_meta('sigungu', years, months)
    record_upload_meta('base',    years, months)
    threading.Thread(target=build_all_agg, args=(ym_list,), daemon=True).start()
    return len(sg_df), len(base_df)


def parse_and_insert_unified(filepath, filter_year=None, filter_month=None):
    df = pd.read_excel(filepath, header=0, engine='openpyxl')
    df.columns = df.columns.astype(str).str.strip()

    date_cols = [c for c in df.columns if re.match(r'^\d{4}\.\d{2}$', str(c).strip())]
    id_cols   = [c for c in df.columns if c not in date_cols]
    if not date_cols:
        raise ValueError("날짜 형식 컬럼(YYYY.MM)을 찾을 수 없습니다.")

    if filter_year and filter_month:
        target = f"{filter_year}.{int(filter_month):02d}"
        if target not in date_cols:
            raise ValueError(f"{target} 컬럼이 없습니다.")
        date_cols = [target]

    long_df = df.melt(id_vars=id_cols, value_vars=date_cols,
                      var_name='yearmonth', value_name='reg_count')
    del df
    long_df = long_df[long_df['reg_count'].notna() & (long_df['reg_count'] > 0)].copy()
    long_df['reg_count'] = long_df['reg_count'].astype(int)
    long_df['year']  = long_df['yearmonth'].str[:4].astype(int)
    long_df['month'] = long_df['yearmonth'].str[5:7].astype(int)
    long_df.drop(columns=['yearmonth'], inplace=True)

    COL_MAP = {
        '지역본부':'headquarters','거점':'base','지역':'region',
        '시군구':'sigungu','법정동':'dong','등록차급':'car_class',
        '수입국가':'import_country','제조사':'maker','등록연료':'fuel','대표차명':'model',
    }
    long_df.rename(columns=COL_MAP, inplace=True)

    for col in ['headquarters','base','region','sigungu','dong',
                'car_class','import_country','maker','fuel','model']:
        long_df[col] = long_df.get(col, pd.Series([''] * len(long_df))) \
                               .fillna('').astype(str).str.strip()

    long_df = long_df[long_df['maker'] != '']
    years   = long_df['year'].unique().tolist()
    months  = long_df['month'].unique().tolist()
    ym_list = [(int(y), int(m)) for y in years for m in months]

    sg_df   = long_df[long_df['sigungu'] != ''].copy()
    base_df = long_df[long_df['base']    != ''].copy()
    del long_df

    con = sqlite3.connect(DB_PATH)
    try:
        for y, m in ym_list:
            con.execute('DELETE FROM car_reg      WHERE year=? AND month=?', (y,m))
            con.execute('DELETE FROM car_reg_base WHERE year=? AND month=?', (y,m))
        con.commit()
        chunk = 5000
        for i in range(0, len(sg_df), chunk):
            sg_df.iloc[i:i+chunk].to_sql('car_reg', con, if_exists='append', index=False)
            con.commit()
        for i in range(0, len(base_df), chunk):
            base_df.iloc[i:i+chunk].to_sql('car_reg_base', con, if_exists='append', index=False)
            con.commit()
    finally:
        con.close()

    record_upload_meta('sigungu', years, months)
    record_upload_meta('base',    years, months)
    threading.Thread(target=build_all_agg, args=(ym_list,), daemon=True).start()
    return len(sg_df), len(base_df)

# ═══════════════════════════════════════════════════════
# WHERE 빌더
# ═══════════════════════════════════════════════════════
def build_where(args, mode='sigungu'):
    clauses, params = [], []
    loc_col = 'sigungu' if mode == 'sigungu' else 'base'
    tbl     = get_tbl_for_mode(mode, args)
    accum   = args.get('accum','') == '1'

    if args.get('year') and args.get('month'):
        clauses.append('year = ?'); params.append(int(args['year']))
        if 'car_reg' in tbl and accum:
            clauses.append('month <= ?')
        else:
            clauses.append('month = ?')
        params.append(int(args['month']))

    hq = args.get('headquarters','').strip()
    if hq and hq != '전체':
        clauses.append('headquarters = ?'); params.append(hq)

    loc_val = args.get(loc_col,'').strip()
    if loc_val and loc_val != '전체':
        add_loc_clause(clauses, params, loc_col, loc_val)

    if mode == 'base':
        for fld in ('base_sigungu','base_dong'):
            val = args.get(fld,'').strip()
            if val and val != '전체':
                col = 'sigungu' if fld == 'base_sigungu' else 'dong'
                clauses.append(f'{col} = ?'); params.append(val)

    if mode == 'sigungu':
        dong = args.get('dong','').strip()
        if dong and dong != '전체':
            clauses.append('dong = ?'); params.append(dong)

    fuel_group = args.get('fuel_group','').strip()
    if fuel_group and fuel_group != '전체':
        clauses.extend(fuel_group_clause(fuel_group))
    else:
        fuel = args.get('fuel','').strip()
        if fuel and fuel != '전체':
            clauses.append('fuel = ?'); params.append(fuel)

    for col in ('car_class','maker'):
        val = args.get(col,'').strip()
        if val and val != '전체':
            clauses.append(f'{col} = ?'); params.append(val)

    return ('WHERE ' + ' AND '.join(clauses)) if clauses else '', params


def build_where_foreign(args, mode='sigungu'):
    clauses, params = [], []
    loc_col = 'sigungu' if mode == 'sigungu' else 'base'
    tbl     = get_tbl_for_mode(mode, args)
    accum   = args.get('accum','') == '1'

    if args.get('year') and args.get('month'):
        clauses.append('year = ?'); params.append(int(args['year']))
        if 'car_reg' in tbl and accum:
            clauses.append('month <= ?')
        else:
            clauses.append('month = ?')
        params.append(int(args['month']))

    hq = args.get('headquarters','').strip()
    if hq and hq != '전체':
        clauses.append('headquarters = ?'); params.append(hq)

    loc_val = args.get(loc_col,'').strip()
    if loc_val and loc_val != '전체':
        add_loc_clause(clauses, params, loc_col, loc_val)

    fuel_group = args.get('fuel_group','').strip()
    if fuel_group and fuel_group != '전체':
        clauses.extend(fuel_group_clause(fuel_group))

    val = args.get('car_class','').strip()
    if val and val != '전체':
        clauses.append('car_class = ?'); params.append(val)

    clauses.append("import_country != '대한민국'")
    clauses.append("maker NOT IN ('현대','기아','TESLA','BYD','르노삼성')")
    return ('WHERE ' + ' AND '.join(clauses)) if clauses else '', params

# ═══════════════════════════════════════════════════════
# M/S 집계
# ═══════════════════════════════════════════════════════
def ms_aggregate(year, month, loc=None, mode='sigungu', accum=False,
                 base_sg=None, base_dong=None, hq=None):
    tbl     = get_agg_table(mode, accum)
    loc_col = 'sigungu' if mode == 'sigungu' else 'base'
    parts, params = ['year=?','month=?'], [year, month]

    if hq and hq != '전체':
        parts.append('headquarters=?'); params.append(hq)
    if loc and loc != '전체':
        if loc == '청주시' and mode == 'sigungu':
            ph = ','.join(['?'] * len(CHEONGJU_DISTS))
            parts.append(f'{loc_col} IN ({ph})')
            params.extend(sorted(CHEONGJU_DISTS))
        else:
            parts.append(f'{loc_col}=?'); params.append(loc)

    where = 'WHERE ' + ' AND '.join(parts)
    rows  = query_db(
        f'SELECT car_class,maker,fuel,SUM(reg_count) AS cnt '
        f'FROM {tbl} {where} GROUP BY car_class,maker,fuel', params)

    result = {}
    for r in rows:
        cg_info = CAR_CLASS_GROUP.get(r['car_class'])
        if not cg_info: continue
        sec, cg = cg_info
        mg      = classify_maker(r['maker'])
        cnt     = r['cnt']
        is_ev   = ('전기' in r['fuel'] and '수소' not in r['fuel']
                   and '하이브리드' not in r['fuel'])
        result.setdefault(sec, {})
        result[sec].setdefault(cg, {})
        result[sec][cg].setdefault(mg, {'cnt':0,'ev':0})
        result[sec][cg][mg]['cnt'] += cnt
        if is_ev: result[sec][cg][mg]['ev'] += cnt

    ordered = {}
    for sec in SECTION_ORDER:
        if sec not in result: continue
        ordered[sec] = {}
        for cls in CLASS_ORDER.get(sec,[]):
            if cls in result[sec]: ordered[sec][cls] = result[sec][cls]
    return ordered

# ═══════════════════════════════════════════════════════
# 엑셀 빌더
# ═══════════════════════════════════════════════════════
def fetch_for_excel(year, month, loc, mode='sigungu', accum=False,
                    base_sg=None, base_dong=None, hq=None):
    use_raw = (mode == 'base' and
               ((base_sg and base_sg != '전체') or
                (base_dong and base_dong != '전체')))

    if use_raw:
        tbl = 'car_reg'
        month_cond = 'month<=?' if accum else 'month=?'
    else:
        tbl = get_agg_table(mode, accum)
        month_cond = 'month=?'

    loc_col = 'sigungu' if mode == 'sigungu' else 'base'
    parts, params = ['year=?', month_cond], [year, month]

    if hq and hq != '전체':
        parts.append('headquarters=?'); params.append(hq)
    if loc and loc != '전체':
        if loc == '청주시' and mode == 'sigungu':
            ph = ','.join(['?'] * len(CHEONGJU_DISTS))
            parts.append(f'{loc_col} IN ({ph})')
            params.extend(sorted(CHEONGJU_DISTS))
        else:
            parts.append(f'{loc_col}=?'); params.append(loc)
    if use_raw:
        if base_sg and base_sg != '전체':
            parts.append('sigungu=?'); params.append(base_sg)
        if base_dong and base_dong != '전체':
            parts.append('dong=?');    params.append(base_dong)

    where = 'WHERE ' + ' AND '.join(parts)
    rows  = query_db(
        f'SELECT car_class, maker, import_country, model, fuel, SUM(reg_count) AS cnt '
        f'FROM {tbl} {where} GROUP BY car_class, maker, import_country, model, fuel',
        params
    )

    d = {}
    for r in rows:
        cc   = r['car_class']
        fuel = r['fuel']
        is_ev = ('전기' in fuel and '수소' not in fuel and '하이브리드' not in fuel)
        # ✅ is_ev를 key에 포함 → EV/비EV 동일 모델 분리
        key = (r['model'], r['maker'], r['import_country'], is_ev)
        d.setdefault(cc, {})
        if key not in d[cc]:
            d[cc][key] = {'cnt': 0, 'is_ev': is_ev}
        d[cc][key]['cnt'] += r['cnt']
    return d


def sum_by_maker(data_dict, orig_classes):
    ALL_COLS  = MAKER_NAMES + ['외산차']
    result    = {mk: 0 for mk in ALL_COLS}
    ev_result = {mk: 0 for mk in ALL_COLS}
    total     = 0
    for cc, models in data_dict.items():
        if orig_classes is not None and cc not in orig_classes: continue
        if not CAR_CLASS_GROUP.get(cc): continue
        # ✅ key = (model, maker, import_country, is_ev)
        for (mo, mk, ic, is_ev), val in models.items():
            cnt = val['cnt']
            col = get_maker_col(mk, ic)
            if col:
                result[col] += cnt
                if is_ev: ev_result[col] += cnt
            total += cnt
    return result, total, ev_result


def build_excel(year, month, loc, mode='sigungu', base_sg=None, base_dong=None, hq=None):
    prev_m_year  = year if month > 1 else year - 1
    prev_m_month = month - 1 if month > 1 else 12

    def mk_fill(h): return PatternFill('solid', fgColor=h)
    def mk_font(bold=False, color='000000', size=9):
        return Font(bold=bold, color=color, size=size, name='맑은 고딕')
    def mk_border():
        s = Side(style='thin', color='BFBFBF')
        return Border(left=s, right=s, top=s, bottom=s)

    BD = mk_border()
    C  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    R  = Alignment(horizontal='right',  vertical='center')
    L  = Alignment(horizontal='left',   vertical='center', indent=1)

    FILL = {
        'header' : mk_fill('1F4E79'),
        'maker'  : mk_fill('2E75B6'),
        'section': mk_fill('4472C4'),
        'cls'    : mk_fill('BDD7EE'),
        'model'  : mk_fill('F2F2F2'),
        'total'  : mk_fill('FFF2CC'),
        'ev'     : mk_fill('E2EFDA'),
    }
    FONTA = {
        'header' : mk_font(True,  'FFFFFF', 9),
        'maker'  : mk_font(True,  'FFFFFF', 9),
        'section': mk_font(True,  'FFFFFF', 10),
        'cls'    : mk_font(True,  '1F4E79', 9),
        'model'  : mk_font(False, '000000', 9),
        'total'  : mk_font(True,  '7F6000', 10),
        'up'     : mk_font(False, 'C00000', 9),
        'down'   : mk_font(False, '1F4E79', 9),
        'ev'     : mk_font(True,  '375623', 10),
    }

    SHOW_MAKERS = {'현대','기아','TESLA','BYD','르노삼성'}
    ALL_MC = MAKER_NAMES + ['외산차']
    FIXED, M_COLS, N = 2, 5, len(ALL_MC)
    def mcs(mi): return FIXED + mi * M_COLS + 1
    ind_s = FIXED + N * M_COLS + 1

    SHEETS = [
        dict(name='전월비교', prev_y=prev_m_year,  prev_m=prev_m_month, accum=False,
             prev_label=f"{str(prev_m_year)[-2:]}년 {prev_m_month}월",
             curr_label=f"{str(year)[-2:]}년 {month}월"),
        dict(name='전년비교', prev_y=year-1, prev_m=month, accum=False,
             prev_label=f"{str(year-1)[-2:]}년 {month}월",
             curr_label=f"{str(year)[-2:]}년 {month}월"),
        dict(name='누계비교', prev_y=year-1, prev_m=month, accum=True,
             prev_label=f"{str(year-1)[-2:]}년 1~{month}월",
             curr_label=f"{str(year)[-2:]}년 1~{month}월"),
    ]

    wb = Workbook()
    wb.remove(wb.active)

    for sd in SHEETS:
        ws = wb.create_sheet(title=sd['name'])

        prev_d = fetch_for_excel(sd['prev_y'], sd['prev_m'], loc, mode,
                                  sd['accum'], base_sg, base_dong, hq)
        curr_d = fetch_for_excel(year, month, loc, mode,
                                  sd['accum'], base_sg, base_dong, hq)

        _, gp, _ = sum_by_maker(prev_d, None); gp = gp or 1
        _, gc, _ = sum_by_maker(curr_d, None); gc = gc or 1

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 20
        for mi in range(N):
            cs = mcs(mi)
            for j, w in enumerate([8,6,8,6,7]):
                ws.column_dimensions[get_column_letter(cs+j)].width = w
        for j, w in enumerate([8,8,7]):
            ws.column_dimensions[get_column_letter(ind_s+j)].width = w

        def sc(row, col, val, fill=None, fnt=None, aln=None):
            c = ws.cell(row=row, column=col, value=val)
            if fill: c.fill = fill
            if fnt:  c.font = fnt
            if aln:  c.alignment = aln
            c.border = BD
            return c

        def merge(r1,c1,r2,c2,val,fill=None,fnt=None,aln=C):
            ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
            cell = ws.cell(row=r1, column=c1, value=val)
            if fill: cell.fill = fill
            if fnt:  cell.font = fnt
            cell.alignment = aln
            for r in range(r1,r2+1):
                for c in range(c1,c2+1):
                    ws.cell(row=r,column=c).border = BD

        # 헤더
        merge(1,1,3,1,'브랜드', FILL['header'],FONTA['header'])
        merge(1,2,3,2,'차종',   FILL['header'],FONTA['header'])
        for mi, mk in enumerate(ALL_MC):
            cs = mcs(mi)
            merge(1,cs,1,cs+M_COLS-1, mk, FILL['maker'],FONTA['maker'])
            merge(2,cs,2,cs+1, sd['prev_label'], FILL['header'],FONTA['header'])
            merge(2,cs+2,2,cs+3, sd['curr_label'], FILL['header'],FONTA['header'])
            sc(2,cs+4,'',FILL['header'],FONTA['header'],C)
            for j,lbl in enumerate(['출고','점유율','출고','점유율','증감율']):
                sc(3,cs+j,lbl,FILL['header'],FONTA['header'],C)
        merge(1,ind_s,1,ind_s+2,'산업수요',FILL['maker'],FONTA['maker'])
        sc(2,ind_s,   sd['prev_label'],FILL['header'],FONTA['header'],C)
        sc(2,ind_s+1, sd['curr_label'],FILL['header'],FONTA['header'],C)
        sc(2,ind_s+2, '',FILL['header'],FONTA['header'],C)
        for j,lbl in enumerate(['출고','출고','증감율']):
            sc(3,ind_s+j,lbl,FILL['header'],FONTA['header'],C)
        for rh in [1,2,3]:
            ws.row_dimensions[rh].height = 18
        ws.freeze_panes = 'A4'

        def write_row(rn, l1, l2, p_mk, c_mk, p_tot, c_tot, rf, ff,
                      la=C, lb=L, denom_p=None, denom_c=None):
            dp = denom_p if denom_p else gp
            dc = denom_c if denom_c else gc
            sc(rn,1,l1,rf,ff,la); sc(rn,2,l2,rf,ff,lb)
            for mi, mk in enumerate(ALL_MC):
                cs  = mcs(mi)
                pc  = p_mk.get(mk,0)
                ccv = c_mk.get(mk,0)
                ms_p = round(pc /dp*100,1) if dp else 0
                ms_c = round(ccv/dc*100,1) if dc else 0
                chg  = round((ccv-pc)/pc*100,1) if pc else None
                sc(rn,cs,   pc or '',           rf,ff,R)
                sc(rn,cs+1, ms_p if pc else '',  rf,ff,R)
                sc(rn,cs+2, ccv or '',           rf,ff,R)
                sc(rn,cs+3, ms_c if ccv else '', rf,ff,R)
                cv  = round(chg,1) if chg is not None else ''
                cc2 = sc(rn,cs+4,cv,rf,ff,R)
                if isinstance(cv,float):
                    cc2.font = FONTA['up'] if cv>0 else FONTA['down']
            ip   = p_tot
            icv  = c_tot
            ichg = round((icv-ip)/ip*100,1) if ip else None
            sc(rn,ind_s,   ip or '',  rf,ff,R)
            sc(rn,ind_s+1, icv or '', rf,ff,R)
            iv  = round(ichg,1) if ichg is not None else ''
            ic2 = sc(rn,ind_s+2,iv,rf,ff,R)
            if isinstance(iv,float):
                ic2.font = FONTA['up'] if iv>0 else FONTA['down']
            ws.row_dimensions[rn].height = 16

        # 본문
        rn = 4
        gp2={mk:0 for mk in ALL_MC}; gc2={mk:0 for mk in ALL_MC}
        gp2_ev={mk:0 for mk in ALL_MC}; gc2_ev={mk:0 for mk in ALL_MC}
        gpt=0; gct=0

        for sec in SECTION_ORDER:
            orig_sec = get_orig_classes(sec=sec)
            psec, pts, psec_ev = sum_by_maker(prev_d, orig_sec)
            csec, cts, csec_ev = sum_by_maker(curr_d, orig_sec)
            lbl = {'승용':'승  용  계','RV':'R  V  계','상용':'상  용  계'}[sec]
            write_row(rn, lbl, '', psec, csec, pts, cts,
                      FILL['section'], FONTA['section'], C, C)
            rn += 1

            for cls in CLASS_ORDER.get(sec,[]):
                orig_cls = get_orig_classes(sec=sec, cls=cls)
                pc2, pt2, _ = sum_by_maker(prev_d, orig_cls)
                cc3, ct2, _ = sum_by_maker(curr_d, orig_cls)
                if pt2==0 and ct2==0: continue
                write_row(rn, cls, '[소계]', pc2, cc3, pt2, ct2,
                          FILL['cls'], FONTA['cls'], C, L)
                rn += 1

                all_keys = set()
                for occ in orig_cls:
                    all_keys |= set(prev_d.get(occ,{}).keys())
                    all_keys |= set(curr_d.get(occ,{}).keys())

                def brand_sort_key(k):
                    mo, mk, ic, is_ev = k  # ✅ is_ev 추가
                    brand = get_brand_label(mk, ic)
                    prefix = ('0' if brand=='현대' else '1' if brand=='기아' else '2_'+brand)
                    cnt = sum(curr_d.get(o,{}).get(k,{}).get('cnt',0) for o in orig_cls)
                    return (prefix, -cnt)

                for (mo, mk, ic, is_ev) in sorted(all_keys, key=brand_sort_key):  # ✅
                    pc3 = sum(prev_d.get(o,{}).get((mo,mk,ic,is_ev),{}).get('cnt',0) for o in orig_cls)
                    cc4 = sum(curr_d.get(o,{}).get((mo,mk,ic,is_ev),{}).get('cnt',0) for o in orig_cls)
                    if pc3==0 and cc4==0: continue
                    col = get_maker_col(mk, ic)
                    if not col: continue
                    rp={m:0 for m in ALL_MC}; rc={m:0 for m in ALL_MC}
                    rp[col]=pc3; rc[col]=cc4
                    write_row(rn, get_brand_label(mk,ic), f'  {mo}',
                              rp, rc, pc3, cc4,
                              FILL['model'], FONTA['model'], C, L,
                              denom_p=pt2, denom_c=ct2)
                    if mk not in SHOW_MAKERS:
                        ws.row_dimensions[rn].hidden = True
                    rn += 1

            for mk in ALL_MC:
                gp2[mk]+=psec.get(mk,0); gc2[mk]+=csec.get(mk,0)
                gp2_ev[mk]+=psec_ev.get(mk,0); gc2_ev[mk]+=csec_ev.get(mk,0)
            gpt+=pts; gct+=cts

        write_row(rn,'전  차  종','',gp2,gc2,gpt,gct,FILL['total'],FONTA['total'],C,C)
        rn += 1

        # EV 행
        gpt_ev = sum(gp2_ev.values())
        gct_ev = sum(gc2_ev.values())
        sc(rn,1,'전기차(EV)',FILL['ev'],FONTA['ev'],C)
        sc(rn,2,'',          FILL['ev'],FONTA['ev'],C)
        for mi, mk in enumerate(ALL_MC):
            cs  = mcs(mi)
            pc  = gp2_ev.get(mk,0)
            ccv = gc2_ev.get(mk,0)
            ms_p = round(pc /gpt_ev*100,1) if gpt_ev else 0
            ms_c = round(ccv/gct_ev*100,1) if gct_ev else 0
            chg  = round((ccv-pc)/pc*100,1) if pc else None
            sc(rn,cs,   pc or '',           FILL['ev'],FONTA['ev'],R)
            sc(rn,cs+1, ms_p if pc else '',  FILL['ev'],FONTA['ev'],R)
            sc(rn,cs+2, ccv or '',           FILL['ev'],FONTA['ev'],R)
            sc(rn,cs+3, ms_c if ccv else '', FILL['ev'],FONTA['ev'],R)
            cv  = round(chg,1) if chg is not None else ''
            cc2 = sc(rn,cs+4,cv,FILL['ev'],FONTA['ev'],R)
            if isinstance(cv,float):
                cc2.font = FONTA['up'] if cv>0 else FONTA['down']
        ichg_ev = round((gct_ev-gpt_ev)/gpt_ev*100,1) if gpt_ev else None
        sc(rn,ind_s,   gpt_ev or '',FILL['ev'],FONTA['ev'],R)
        sc(rn,ind_s+1, gct_ev or '',FILL['ev'],FONTA['ev'],R)
        iv_ev  = round(ichg_ev,1) if ichg_ev is not None else ''
        ic2_ev = sc(rn,ind_s+2,iv_ev,FILL['ev'],FONTA['ev'],R)
        if isinstance(iv_ev,float):
            ic2_ev.font = FONTA['up'] if iv_ev>0 else FONTA['down']
        ws.row_dimensions[rn].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ═══════════════════════════════════════════════════════
# 라우트
# ═══════════════════════════════════════════════════════
@app.route('/')
def index(): return render_template('index.html')

@app.route('/upload/login', methods=['GET','POST'])
def upload_login():
    error = ''
    if request.method == 'POST':
        if request.form.get('password','') == UPLOAD_PASSWORD:
            session['upload_auth'] = True
            return redirect(url_for('upload_page'))
        error = '비밀번호가 올바르지 않습니다.'
    return render_template('upload_login.html', error=error)

@app.route('/upload/logout')
def upload_logout():
    session.pop('upload_auth', None)
    return redirect(url_for('upload_login'))

@app.route('/upload')
def upload_page():
    if not session.get('upload_auth'):
        return redirect(url_for('upload_login'))
    return render_template('upload.html')


@app.route('/api/upload', methods=['POST'])
def upload():
    if not session.get('upload_auth'):
        return jsonify({'ok':False,'msg':'인증이 필요합니다.'}), 401
    if 'file' not in request.files:
        return jsonify({'ok':False,'msg':'파일이 없습니다.'}), 400
    f = request.files['file']
    if not f.filename.endswith(('.xlsx','.xls')):
        return jsonify({'ok':False,'msg':'엑셀 파일만 가능합니다.'}), 400

    upload_type  = request.form.get('type','unified')
    filter_year  = request.form.get('year','').strip()
    filter_month = request.form.get('month','').strip()
    save_path = '/tmp/uploaded_temp.xlsx'
    f.save(save_path)
    try:
        fy    = int(filter_year)  if filter_year  else None
        fm    = int(filter_month) if filter_month else None
        label = f"{fy}년 {fm}월" if fy and fm else "전체 기간"
        if upload_type == 'national':
            cnt = parse_and_insert_national(save_path, fy, fm)
            return jsonify({'ok':True,'msg':f'[전국/{label}] {cnt:,}건 적재 완료 (집계 중...)'})
        else:
            sg_cnt, base_cnt = parse_and_insert_unified(save_path, fy, fm)
            return jsonify({'ok':True,
                            'msg':f'[{label}] 시군구 {sg_cnt:,}건 / 거점 {base_cnt:,}건 적재 완료 (집계 중...)'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'ok':False,'msg':str(e)}), 500


@app.route('/api/agg_status')
def agg_status():
    return jsonify({'ready': _agg_ready})


@app.route('/api/rebuild_agg', methods=['POST'])
def rebuild_agg():
    if not session.get('upload_auth'):
        return jsonify({'ok':False,'msg':'인증 필요'}), 401
    global _agg_ready
    _agg_ready = False
    threading.Thread(target=build_all_agg, daemon=True).start()
    return jsonify({'ok':True,'msg':'백그라운드 재집계 시작됨'})


@app.route('/api/delete_all', methods=['POST'])
def delete_all():
    if not session.get('upload_auth'):
        return jsonify({'ok':False,'msg':'인증이 필요합니다.'}), 401
    con = get_con()
    for tbl in ['car_reg','car_reg_base',
                'agg_sigungu','agg_base',
                'agg_sigungu_accum','agg_base_accum']:
        con.execute(f'DELETE FROM {tbl}')
    con.execute("DELETE FROM upload_meta WHERE mode!='national'")
    con.commit()
    return jsonify({'ok':True,'msg':'충북 데이터 전체 삭제 완료'})


@app.route('/api/delete_national', methods=['POST'])
def delete_national():
    if not session.get('upload_auth'):
        return jsonify({'ok':False,'msg':'인증이 필요합니다.'}), 401
    con = get_con()
    for tbl in ['national_reg','agg_national','agg_national_accum']:
        con.execute(f'DELETE FROM {tbl}')
    con.execute("DELETE FROM upload_meta WHERE mode='national'")
    con.commit()
    _nat_cache.clear(); _nat_cache_time.clear()
    return jsonify({'ok':True,'msg':'전국 데이터 전체 삭제 완료'})


@app.route('/api/options')
def options():
    mode    = request.args.get('mode','sigungu')
    tbl     = get_agg_table(mode, False)
    loc_col = 'sigungu' if mode == 'sigungu' else 'base'
    sigungu = request.args.get('sigungu','').strip()
    base    = request.args.get('base','').strip()
    base_sg = request.args.get('base_sigungu','').strip()
    hq      = request.args.get('headquarters','').strip()
    hq_cond = f"AND headquarters='{hq}'" if hq and hq != '전체' else ''

    hq_list = [r['headquarters'] for r in query_db(
        f'SELECT DISTINCT headquarters FROM {tbl} WHERE headquarters!="" ORDER BY headquarters')]

    locs = [r[loc_col] for r in query_db(
        f'SELECT DISTINCT {loc_col} FROM {tbl} WHERE {loc_col}!="" {hq_cond} ORDER BY {loc_col}')]
    if mode == 'sigungu':
        cheongju_subs = [l for l in locs if l.startswith('청주시 ')]
        if cheongju_subs:
            idx = locs.index(cheongju_subs[0])
            locs.insert(idx,'청주시')

    dongs = []
    if mode == 'sigungu' and sigungu and sigungu != '전체':
        if sigungu == '청주시':
            ph = ','.join(['?'] * len(CHEONGJU_DISTS))
            dongs = [r['dong'] for r in query_db(
                f"SELECT DISTINCT dong FROM car_reg WHERE sigungu IN ({ph}) AND dong!='' {hq_cond} ORDER BY dong",
                list(sorted(CHEONGJU_DISTS)))]
        else:
            dongs = [r['dong'] for r in query_db(
                f"SELECT DISTINCT dong FROM car_reg WHERE sigungu=? AND dong!='' {hq_cond} ORDER BY dong",
                [sigungu])]

    base_sgs = []
    if mode == 'base' and base and base != '전체':
        base_sgs = [r['sigungu'] for r in query_db(
            f"SELECT DISTINCT sigungu FROM car_reg WHERE base=? AND sigungu!='' {hq_cond} ORDER BY sigungu",
            [base])]

    base_dongs = []
    if mode == 'base' and base_sg and base_sg != '전체':
        base_dongs = [r['dong'] for r in query_db(
            f"SELECT DISTINCT dong FROM car_reg WHERE sigungu=? AND dong!='' {hq_cond} ORDER BY dong",
            [base_sg])]

    ym_list = query_db(
        f'SELECT DISTINCT year,month FROM {tbl} WHERE 1=1 {hq_cond} ORDER BY year,month')

    return jsonify({
        'years'       : sorted(set(r['year']  for r in ym_list)),
        'months'      : list(range(1,13)),
        'headquarters': hq_list,
        'locations'   : locs,
        'dongs'       : dongs,
        'base_sgs'    : base_sgs,
        'base_dongs'  : base_dongs,
        'car_classes' : [r['car_class'] for r in query_db(
                            f'SELECT DISTINCT car_class FROM {tbl} ORDER BY car_class')],
        'makers'      : [r['maker'] for r in query_db(
                            f'SELECT DISTINCT maker FROM {tbl} ORDER BY maker')],
        'ym_list'     : ym_list,
    })


@app.route('/api/summary')
def summary():
    mode  = request.args.get('mode','sigungu')
    tbl   = get_tbl_for_mode(mode, request.args)
    where, params = build_where(request.args, mode)
    accum = request.args.get('accum','') == '1'

    def total(extra=None):
        sql = f'SELECT COALESCE(SUM(reg_count),0) AS v FROM {tbl} {where}'
        if extra: sql += ' AND ' + ' AND '.join(extra)
        return query_db(sql, params)[0]['v']

    t       = total()
    hyundai = total(["maker='현대'"])
    kia     = total(["maker='기아'"])
    ms_hy   = round(hyundai/t*100,1) if t else 0
    ms_kia  = round(kia    /t*100,1) if t else 0

    args_no_fuel = {k:v for k,v in request.args.items() if k not in ('fuel_group','fuel')}
    where_b, params_b = build_where(args_no_fuel, mode)

    def total_b(extra=None):
        sql = f'SELECT COALESCE(SUM(reg_count),0) AS v FROM {tbl} {where_b}'
        if extra: sql += ' AND ' + ' AND '.join(extra)
        return query_db(sql, params_b)[0]['v']

    ice_c     = ["fuel NOT LIKE '%전기%'","fuel NOT LIKE '%수소%'"]
    ev_c      = ["(fuel LIKE '%전기%' OR fuel LIKE '%수소%')"]
    ice_total = total_b(ice_c); ice_hk = total_b(ice_c+["maker='현대'"])
    ev_total  = total_b(ev_c);  ev_hk  = total_b(ev_c +["maker='현대'"])

    year  = request.args.get('year', type=int)
    month = request.args.get('month',type=int)
    nat_ms, _ = get_national_ms(year, month, accum)

    return jsonify({
        'total':t,'hyundai':hyundai,'kia':kia,
        'ms_hyundai':ms_hy,'ms_kia':ms_kia,
        'ice_total':ice_total,'ice_hk':ice_hk,
        'ev_total':ev_total,'ev_hk':ev_hk,
        'nat_ms':nat_ms,
    })


@app.route('/api/by_maker')
def by_maker():
    mode    = request.args.get('mode','sigungu')
    tbl     = get_tbl_for_mode(mode, request.args)
    where, params = build_where(request.args, mode)
    accum   = request.args.get('accum','') == '1'

    rows = query_db(
        f'SELECT maker,SUM(reg_count) AS total FROM {tbl} {where} '
        f'GROUP BY maker ORDER BY total DESC LIMIT 15', params)
    grand = query_db(
        f'SELECT COALESCE(SUM(reg_count),0) AS v FROM {tbl} {where}', params)[0]['v'] or 1
    for r in rows: r['ms'] = round(r['total']/grand*100,1)

    loc_col   = 'sigungu' if mode == 'sigungu' else 'base'
    loc_val   = request.args.get(loc_col,'').strip()
    base_sg   = request.args.get('base_sigungu','').strip()
    base_dong = request.args.get('base_dong','').strip()
    is_filtered = bool((loc_val and loc_val != '전체') or
                       (base_sg and base_sg != '전체') or
                       (base_dong and base_dong != '전체'))
    chb_map = {}
    if is_filtered:
        args_no_loc = {k:v for k,v in request.args.items()
                       if k not in ('sigungu','base','dong','base_sigungu','base_dong')}
        tbl_chb = get_tbl_for_mode(mode, args_no_loc)
        where_chb, params_chb = build_where(args_no_loc, mode)
        grand_chb = query_db(
            f'SELECT COALESCE(SUM(reg_count),0) AS v FROM {tbl_chb} {where_chb}',
            params_chb)[0]['v'] or 1
        chb_rows = query_db(
            f'SELECT maker,SUM(reg_count) AS total FROM {tbl_chb} {where_chb} GROUP BY maker',
            params_chb)
        chb_map = {r['maker']:round(r['total']/grand_chb*100,1) for r in chb_rows}

    year  = request.args.get('year', type=int)
    month = request.args.get('month',type=int)
    nat_ms_map, _ = get_national_ms(year, month, accum)

    return jsonify({'rows':rows,'grand':grand,
                    'chb_map':chb_map,'is_filtered':is_filtered,
                    'nat_ms_map':nat_ms_map})


@app.route('/api/by_model')
def by_model():
    mode    = request.args.get('mode','sigungu')
    foreign = request.args.get('foreign','')
    tbl     = get_tbl_for_mode(mode, request.args)
    where, params = (build_where_foreign(request.args, mode)
                     if foreign == '1' else build_where(request.args, mode))
    return jsonify(query_db(
        f'SELECT maker,model,SUM(reg_count) AS total FROM {tbl} {where} '
        f'GROUP BY maker,model ORDER BY total DESC LIMIT 10', params))


@app.route('/api/db_status')
def db_status():
    mode = request.args.get('mode','sigungu')
    tbl  = get_agg_table(mode, False)
    return jsonify(query_db(
        f'SELECT year,month,COUNT(*) AS rows,SUM(reg_count) AS total '
        f'FROM {tbl} GROUP BY year,month ORDER BY year,month'))


@app.route('/api/memo', methods=['GET'])
def get_memo():
    mode = request.args.get('mode','sigungu')
    rows = query_db('SELECT memo,saved_at FROM data_memo WHERE mode=?',(mode,))
    if rows: return jsonify({'ok':True,'memo':rows[0]['memo'],'saved_at':rows[0]['saved_at']})
    return jsonify({'ok':False,'memo':''})

@app.route('/api/memo', methods=['POST'])
def save_memo():
    mode = request.args.get('mode','sigungu')
    memo = request.json.get('memo','').strip()
    now  = datetime.now().strftime('%Y-%m-%d %H:%M')
    con  = get_con()
    con.execute('DELETE FROM data_memo WHERE mode=?',(mode,))
    con.execute('INSERT INTO data_memo(mode,memo,saved_at) VALUES(?,?,?)',(mode,memo,now))
    con.commit()
    return jsonify({'ok':True,'saved_at':now})


@app.route('/api/ms_report')
def ms_report():
    year      = request.args.get('year',        type=int)
    month     = request.args.get('month',       type=int)
    mode      = request.args.get('mode',        'sigungu')
    loc       = request.args.get('loc',         '전체')
    base_sg   = request.args.get('base_sigungu','').strip()
    base_dong = request.args.get('base_dong',   '').strip()
    hq        = request.args.get('headquarters','').strip()
    if not year or not month:
        return jsonify({'ok':False,'msg':'year, month 필수'}), 400

    pmy = year if month > 1 else year-1
    pmm = month-1 if month > 1 else 12

    def mk(y,m,acc=False):
        return ms_aggregate(y,m,loc,mode,acc,base_sg or None,base_dong or None,hq or None)
    def at(d): return sum(mg['cnt'] for s in d.values() for c in s.values() for mg in c.values())
    def ae(d): return sum(mg['ev']  for s in d.values() for c in s.values() for mg in c.values())

    curr = mk(year,month)
    return jsonify({
        'year':year,'month':month,'loc':loc,'mode':mode,
        'mom':{'curr':curr,'prev':mk(pmy,pmm),
               'curr_total':at(curr),'prev_total':at(mk(pmy,pmm)),
               'curr_ev':ae(curr),'prev_ev':ae(mk(pmy,pmm)),
               'prev_label':f"{str(pmy)[-2:]}년 {pmm}월",
               'curr_label':f"{str(year)[-2:]}년 {month}월"},
        'yoy':{'curr':curr,'prev':mk(year-1,month),
               'curr_total':at(curr),'prev_total':at(mk(year-1,month)),
               'curr_ev':ae(curr),'prev_ev':ae(mk(year-1,month)),
               'prev_label':f"{str(year-1)[-2:]}년 {month}월",
               'curr_label':f"{str(year)[-2:]}년 {month}월"},
        'accum':{'curr':mk(year,month,True),'prev':mk(year-1,month,True),
                 'curr_total':at(mk(year,month,True)),'prev_total':at(mk(year-1,month,True)),
                 'curr_ev':ae(mk(year,month,True)),'prev_ev':ae(mk(year-1,month,True)),
                 'prev_label':f"{str(year-1)[-2:]}년 1~{month}월",
                 'curr_label':f"{str(year)[-2:]}년 1~{month}월"},
    })


@app.route('/api/ms_report/download')
def ms_report_download():
    year      = request.args.get('year',        type=int)
    month     = request.args.get('month',       type=int)
    mode      = request.args.get('mode',        'sigungu')
    loc       = request.args.get('loc',         '전체')
    base_sg   = request.args.get('base_sigungu','').strip()
    base_dong = request.args.get('base_dong',   '').strip()
    hq        = request.args.get('headquarters','').strip()
    if not year or not month:
        return jsonify({'ok':False,'msg':'year, month 필수'}), 400
    buf      = build_excel(year,month,loc,mode,base_sg or None,base_dong or None,hq or None)
    loc_lbl  = loc if loc and loc != '전체' else '전체'
    mode_lbl = '거점' if mode == 'base' else '시군구'
    hq_lbl   = hq if hq else '전체'
    filename = f"MS_리포트_{hq_lbl}_{mode_lbl}_{loc_lbl}_{year}년{month:02d}월.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/maker_analysis')
def maker_analysis():
    makers = request.args.getlist('maker') or [request.args.get('maker','').strip()]
    makers = [m for m in makers if m]
    year   = request.args.get('year', type=int)
    month  = request.args.get('month',type=int)
    accum  = request.args.get('accum','') == '1'
    hq     = request.args.get('headquarters','').strip()
    if not makers or not year or not month:
        return jsonify({'ok':False,'msg':'maker, year, month 필수'}), 400

    tbl        = get_agg_table('sigungu', accum)
    month_cond = 'month=?'
    fuel_group   = request.args.get('fuel_group','').strip()
    fuel_clauses = fuel_group_clause(fuel_group) if fuel_group else []
    fuel_sql     = (' AND '+' AND '.join(fuel_clauses)) if fuel_clauses else ''
    car_class  = request.args.get('car_class','').strip()
    cc_sql     = f" AND car_class='{car_class}'" if car_class and car_class != '전체' else ''
    hq_sql     = f" AND headquarters='{hq}'"     if hq and hq != '전체' else ''

    total_sql = (f'SELECT sigungu,SUM(reg_count) AS total FROM {tbl} '
                 f'WHERE year=? AND {month_cond} {hq_sql} {fuel_sql} {cc_sql} GROUP BY sigungu')

    def tmap(y,m): return {r['sigungu']:r['total'] for r in query_db(total_sql,[y,m])}
    pmy,pmm = (year,month-1) if month>1 else (year-1,12)
    total_map    = tmap(year,  month)
    total_py_map = tmap(year-1,month)
    total_pm_map = tmap(pmy,   pmm)

    maker_sql = (f'SELECT sigungu,SUM(reg_count) AS total FROM {tbl} '
                 f'WHERE year=? AND {month_cond} AND maker=? {hq_sql} {fuel_sql} {cc_sql} GROUP BY sigungu')
    chb_total    = sum(total_map.values())    or 1
    chb_total_py = sum(total_py_map.values()) or 1
    chb_total_pm = sum(total_pm_map.values()) or 1
    nat_ms_map,_ = get_national_ms(year,month,accum)

    makers_data = {}
    for mk in makers:
        curr_map = {r['sigungu']:r['total'] for r in query_db(maker_sql,[year,  month,mk])}
        py_map   = {r['sigungu']:r['total'] for r in query_db(maker_sql,[year-1,month,mk])}
        pm_map   = {r['sigungu']:r['total'] for r in query_db(maker_sql,[pmy,   pmm,  mk])}
        chb_maker = sum(curr_map.values())
        chb_ms    = round(chb_maker/chb_total*100,1)
        chb_ms_py = round(sum(py_map.values())/chb_total_py*100,1)
        chb_ms_pm = round(sum(pm_map.values())/chb_total_pm*100,1)
        nat_ms    = nat_ms_map.get(mk)
        makers_data[mk] = {
            'curr_map':curr_map,'py_map':py_map,'pm_map':pm_map,
            'chb_maker':chb_maker,'chb_ms':chb_ms,
            'chb_ms_py':chb_ms_py,'chb_ms_pm':chb_ms_pm,
            'yoy_p':round(chb_ms-chb_ms_py,1),'mom_p':round(chb_ms-chb_ms_pm,1),
            'nat_ms':nat_ms,
            'nat_diff':round(chb_ms-nat_ms,1) if nat_ms is not None else None,
        }

    all_locs = sorted(set(list(total_map.keys()) +
                          [l for md in makers_data.values() for l in md['curr_map'].keys()]))
    CHEONGJU_SUBS = {'청주시 상당구','청주시 서원구','청주시 청원구','청주시 흥덕구'}
    if any(l in CHEONGJU_SUBS for l in all_locs):
        cj_locs    = [l for l in all_locs if l in CHEONGJU_SUBS]
        other_locs = [l for l in all_locs if l not in CHEONGJU_SUBS and not l.startswith('청주시')]
        for mp in [total_map,total_py_map,total_pm_map]:
            mp['청주시'] = sum(mp.get(l,0) for l in cj_locs) or 1
        for md in makers_data.values():
            for key in ['curr_map','py_map','pm_map']:
                md[key]['청주시'] = sum(md[key].get(l,0) for l in cj_locs)
        all_locs_final = sorted(other_locs) + ['청주시'] + sorted(cj_locs)
    else:
        all_locs_final = [l for l in all_locs if l]
        CHEONGJU_SUBS  = set()

    rows = []
    for loc in all_locs_final:
        if not loc: continue
        t=total_map.get(loc,0) or 1
        tpy=total_py_map.get(loc,0) or 1
        tpm=total_pm_map.get(loc,0) or 1
        maker_stats = {}
        for mk,md in makers_data.items():
            m=md['curr_map'].get(loc,0); mpy=md['py_map'].get(loc,0); mpm=md['pm_map'].get(loc,0)
            ms=round(m/t*100,1); ms_py=round(mpy/tpy*100,1); ms_pm=round(mpm/tpm*100,1)
            nat_ms=md['nat_ms']
            maker_stats[mk]={'total':m,'ms':ms,'yoy_p':round(ms-ms_py,1),
                             'mom_p':round(ms-ms_pm,1),'vs_chb':round(ms-md['chb_ms'],1),
                             'vs_nat':round(ms-nat_ms,1) if nat_ms is not None else None}
        rows.append({'loc':loc,'total':total_map.get(loc,0),
                     'is_sub':loc in CHEONGJU_SUBS,'maker_stats':maker_stats})

    return jsonify({
        'ok':True,'makers':makers,'year':year,'month':month,'accum':accum,
        'chb_total':sum(v for k,v in total_map.items() if k not in CHEONGJU_SUBS),
        'nat_ms_map':nat_ms_map,
        'makers_data':{mk:{'chb_maker':md['chb_maker'],'chb_ms':md['chb_ms'],
                            'chb_ms_py':md['chb_ms_py'],'chb_ms_pm':md['chb_ms_pm'],
                            'yoy_p':md['yoy_p'],'mom_p':md['mom_p'],
                            'nat_ms':md['nat_ms'],'nat_diff':md['nat_diff']}
                        for mk,md in makers_data.items()},
        'rows':rows,
    })


@app.route('/api/maker_analysis_base')
def maker_analysis_base():
    makers = request.args.getlist('maker') or [request.args.get('maker','').strip()]
    makers = [m for m in makers if m]
    year   = request.args.get('year', type=int)
    month  = request.args.get('month',type=int)
    accum  = request.args.get('accum','') == '1'
    hq     = request.args.get('headquarters','').strip()
    if not makers or not year or not month:
        return jsonify({'ok':False,'msg':'maker, year, month 필수'}), 400

    tbl        = get_agg_table('base', accum)
    month_cond = 'month=?'
    fuel_group   = request.args.get('fuel_group','').strip()
    fuel_clauses = fuel_group_clause(fuel_group) if fuel_group else []
    fuel_sql     = (' AND '+' AND '.join(fuel_clauses)) if fuel_clauses else ''
    car_class  = request.args.get('car_class','').strip()
    cc_sql     = f" AND car_class='{car_class}'" if car_class and car_class != '전체' else ''
    hq_sql     = f" AND headquarters='{hq}'"     if hq and hq != '전체' else ''

    total_sql = (f'SELECT base,SUM(reg_count) AS total FROM {tbl} '
                 f'WHERE year=? AND {month_cond} AND base!="" {hq_sql} {fuel_sql} {cc_sql} GROUP BY base')
    pmy,pmm   = (year,month-1) if month>1 else (year-1,12)

    def tmap(y,m): return {r['base']:r['total'] for r in query_db(total_sql,[y,m])}
    total_map    = tmap(year,  month)
    total_py_map = tmap(year-1,month)
    total_pm_map = tmap(pmy,   pmm)

    maker_sql = (f'SELECT base,SUM(reg_count) AS total FROM {tbl} '
                 f'WHERE year=? AND {month_cond} AND maker=? AND base!="" {hq_sql} {fuel_sql} {cc_sql} GROUP BY base')
    chb_total    = sum(total_map.values())    or 1
    chb_total_py = sum(total_py_map.values()) or 1
    chb_total_pm = sum(total_pm_map.values()) or 1
    nat_ms_map,_ = get_national_ms(year,month,accum)

    makers_meta = {}
    for mk in makers:
        curr_map={r['base']:r['total'] for r in query_db(maker_sql,[year,  month,mk])}
        py_map  ={r['base']:r['total'] for r in query_db(maker_sql,[year-1,month,mk])}
        pm_map  ={r['base']:r['total'] for r in query_db(maker_sql,[pmy,   pmm,  mk])}
        chb_maker=sum(curr_map.values())
        chb_ms   =round(chb_maker/chb_total*100,1)
        chb_ms_py=round(sum(py_map.values())/chb_total_py*100,1)
        chb_ms_pm=round(sum(pm_map.values())/chb_total_pm*100,1)
        nat_ms   =nat_ms_map.get(mk)
        makers_meta[mk]={
            'curr_map':curr_map,'py_map':py_map,'pm_map':pm_map,
            'chb_maker':chb_maker,'chb_ms':chb_ms,
            'yoy_p':round(chb_ms-chb_ms_py,1),'mom_p':round(chb_ms-chb_ms_pm,1),
            'nat_ms':nat_ms,'nat_diff':round(chb_ms-nat_ms,1) if nat_ms is not None else None,
        }

    sg_sql = (f'SELECT base,sigungu,SUM(reg_count) AS total FROM car_reg '
              f'WHERE year=? AND {"month<=?" if accum else "month=?"} AND base!="" {hq_sql} {fuel_sql} {cc_sql} GROUP BY base,sigungu')
    mk_sg_sql = (f'SELECT base,sigungu,SUM(reg_count) AS total FROM car_reg '
                 f'WHERE year=? AND {"month<=?" if accum else "month=?"} AND maker=? AND base!="" {hq_sql} {fuel_sql} {cc_sql} GROUP BY base,sigungu')

    sg_rows    = query_db(sg_sql,[year,  month])
    sg_py_rows = query_db(sg_sql,[year-1,month])
    sg_pm_rows = query_db(sg_sql,[pmy,   pmm])
    sg_total_map    = {(r['base'],r['sigungu']):r['total'] for r in sg_rows}
    sg_total_py_map = {(r['base'],r['sigungu']):r['total'] for r in sg_py_rows}
    sg_total_pm_map = {(r['base'],r['sigungu']):r['total'] for r in sg_pm_rows}
    base_sg_map = {}
    for r in sg_rows:
        base_sg_map.setdefault(r['base'],[])
        if r['sigungu'] not in base_sg_map[r['base']]:
            base_sg_map[r['base']].append(r['sigungu'])

    mk_sg_data = {}
    for mk in makers:
        mk_sg_data[mk] = {
            'curr':{(r['base'],r['sigungu']):r['total'] for r in query_db(mk_sg_sql,[year,  month,mk])},
            'py'  :{(r['base'],r['sigungu']):r['total'] for r in query_db(mk_sg_sql,[year-1,month,mk])},
            'pm'  :{(r['base'],r['sigungu']):r['total'] for r in query_db(mk_sg_sql,[pmy,   pmm,  mk])},
        }

    rows = []
    for base in sorted(total_map.keys()):
        bt=total_map.get(base,0) or 1
        bpy=total_py_map.get(base,0) or 1
        bpm=total_pm_map.get(base,0) or 1
        maker_stats={}
        for mk,md in makers_meta.items():
            m=md['curr_map'].get(base,0); mpy=md['py_map'].get(base,0); mpm=md['pm_map'].get(base,0)
            ms=round(m/bt*100,1); ms_py=round(mpy/bpy*100,1); ms_pm=round(mpm/bpm*100,1)
            nat_ms=md['nat_ms']
            maker_stats[mk]={'total':m,'ms':ms,'yoy_p':round(ms-ms_py,1),
                             'mom_p':round(ms-ms_pm,1),'vs_chb':round(ms-md['chb_ms'],1),
                             'vs_nat':round(ms-nat_ms,1) if nat_ms is not None else None}
        detail=[]
        for sg in sorted(base_sg_map.get(base,[])):
            sgt=sg_total_map.get((base,sg),0) or 1
            sgpy=sg_total_py_map.get((base,sg),0) or 1
            sgpm=sg_total_pm_map.get((base,sg),0) or 1
            sg_stats={}
            for mk in makers:
                m=mk_sg_data[mk]['curr'].get((base,sg),0)
                mpy=mk_sg_data[mk]['py'].get((base,sg),0)
                mpm=mk_sg_data[mk]['pm'].get((base,sg),0)
                ms=round(m/sgt*100,1); ms_py=round(mpy/sgpy*100,1); ms_pm=round(mpm/sgpm*100,1)
                nat_ms=makers_meta[mk]['nat_ms']
                sg_stats[mk]={'total':m,'ms':ms,'yoy_p':round(ms-ms_py,1),
                              'mom_p':round(ms-ms_pm,1),'vs_chb':round(ms-makers_meta[mk]['chb_ms'],1),
                              'vs_nat':round(ms-nat_ms,1) if nat_ms is not None else None}
            detail.append({'loc':sg,'total':sg_total_map.get((base,sg),0),'maker_stats':sg_stats})
        rows.append({'base':base,'total':total_map.get(base,0),'maker_stats':maker_stats,'detail':detail})

    return jsonify({
        'ok':True,'makers':makers,'chb_total':chb_total,'nat_ms_map':nat_ms_map,
        'makers_data':{mk:{'chb_maker':md['chb_maker'],'chb_ms':md['chb_ms'],
                            'yoy_p':md['yoy_p'],'mom_p':md['mom_p'],
                            'nat_ms':md['nat_ms'],'nat_diff':md['nat_diff']}
                        for mk,md in makers_meta.items()},
        'rows':rows,
    })


@app.route('/api/maker_models')
def maker_models():
    mode    = request.args.get('mode','sigungu')
    maker   = request.args.get('maker','').strip()
    year    = request.args.get('year', type=int)
    month   = request.args.get('month',type=int)
    accum   = request.args.get('accum','') == '1'
    sigungu = request.args.get('sigungu','').strip()
    hq      = request.args.get('headquarters','').strip()
    if not maker or not year or not month: return jsonify([])

    tbl     = get_agg_table(mode, accum)
    loc_col = 'sigungu' if mode == 'sigungu' else 'base'
    loc_cond, loc_params = '', []
    hq_cond  = f" AND headquarters='{hq}'" if hq and hq != '전체' else ''
    fuel_group   = request.args.get('fuel_group','').strip()
    fuel_clauses = fuel_group_clause(fuel_group) if fuel_group else []
    fuel_cond    = (' AND '+' AND '.join(fuel_clauses)) if fuel_clauses else ''

    if sigungu and sigungu != '전체':
        if sigungu == '청주시' and mode == 'sigungu':
            ph = ','.join(['?']*len(CHEONGJU_DISTS))
            loc_cond=f' AND {loc_col} IN ({ph})'; loc_params=list(sorted(CHEONGJU_DISTS))
        else:
            loc_cond=f' AND {loc_col}=?'; loc_params=[sigungu]

    sql = (f'SELECT model,SUM(reg_count) AS total FROM {tbl} '
           f'WHERE year=? AND month=? AND maker=? {loc_cond} {hq_cond} {fuel_cond} '
           f'GROUP BY model ORDER BY total DESC LIMIT 15')
    curr    = query_db(sql,[year,  month,maker]+loc_params)
    py_rows = query_db(sql,[year-1,month,maker]+loc_params)
    pmy,pmm = (year,month-1) if month>1 else (year-1,12)
    pm_rows = query_db(sql,[pmy,pmm,maker]+loc_params)
    py_map  = {r['model']:r['total'] for r in py_rows}
    pm_map  = {r['model']:r['total'] for r in pm_rows}
    grand   = sum(r['total'] for r in curr) or 1

    return jsonify([{
        'model':r['model'],'total':r['total'],'share':round(r['total']/grand*100,1),
        'prev_yr':py_map.get(r['model'],0),'prev_mon':pm_map.get(r['model'],0),
        'yoy':round((r['total']-py_map.get(r['model'],0))/py_map.get(r['model'],0)*100,1)
              if py_map.get(r['model'],0) else None,
        'mom':round((r['total']-pm_map.get(r['model'],0))/pm_map.get(r['model'],0)*100,1)
              if pm_map.get(r['model'],0) else None,
    } for r in curr])


@app.route('/api/maker_fuel')
def maker_fuel():
    mode    = request.args.get('mode','sigungu')
    maker   = request.args.get('maker','').strip()
    year    = request.args.get('year', type=int)
    month   = request.args.get('month',type=int)
    accum   = request.args.get('accum','') == '1'
    sigungu = request.args.get('sigungu','').strip()
    hq      = request.args.get('headquarters','').strip()
    if not maker or not year or not month: return jsonify([])

    tbl     = get_agg_table(mode, accum)
    loc_col = 'sigungu' if mode == 'sigungu' else 'base'
    loc_cond, loc_params = '', []
    hq_cond = f" AND headquarters='{hq}'" if hq and hq != '전체' else ''

    if sigungu and sigungu != '전체':
        if sigungu == '청주시' and mode == 'sigungu':
            ph = ','.join(['?']*len(CHEONGJU_DISTS))
            loc_cond=f' AND {loc_col} IN ({ph})'; loc_params=list(sorted(CHEONGJU_DISTS))
        else:
            loc_cond=f' AND {loc_col}=?'; loc_params=[sigungu]

    rows = query_db(
        f'SELECT fuel,SUM(reg_count) AS total FROM {tbl} '
        f'WHERE year=? AND month=? AND maker=? {loc_cond} {hq_cond} GROUP BY fuel ORDER BY total DESC',
        [year,month,maker]+loc_params)

    grand  = sum(r['total'] for r in rows) or 1
    groups = {'ICE':0,'HEV':0,'EV':0,'FCEV':0}
    for r in rows:
        f = r['fuel']
        if '수소' in f:                             groups['FCEV'] += r['total']
        elif '전기' in f and '하이브리드' not in f: groups['EV']   += r['total']
        elif '하이브리드' in f:                     groups['HEV']  += r['total']
        else:                                       groups['ICE']  += r['total']

    return jsonify([{'group':k,'total':v,'share':round(v/grand*100,1)} for k,v in groups.items()])


@app.after_request
def no_cache(response):
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma']        = 'no-cache'
    response.headers['Expires']       = '0'
    return response


if __name__ == '__main__':
    init_db()

    def _boot_agg():
        time.sleep(1)
        try:
            build_all_agg()
        except Exception as e:
            print(f'[BOOT-AGG] 오류: {e}')

    threading.Thread(target=_boot_agg, daemon=True).start()
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5500)), debug=False)
